import asyncio
import logging
import reflex as rx
from typing import List, Dict, Any
import csv
import io
from datetime import datetime
from ..database_manager import obtener_conexion
from fpdf import FPDF
from .estado_boveda import EstadoBoveda

logger = logging.getLogger(__name__)

class EstadoReportes(rx.State):
    resumen_global: Dict[str, int] = {
        "total_estudiantes": 0,
        "con_pasantia": 0,
        "sin_pasantia": 0,
        "total_tesis": 0
    }
    estadisticas_carreras: List[Dict[str, Any]] = []
    mejores_tutores: List[Dict[str, Any]] = []
    mejores_empresas: List[Dict[str, Any]] = []
    procesando: bool = False

    async def exportar_tesis_excel(self):
        """
        Genera y descarga un Excel con todos los Trabajos de Grado registrados de manera profesional.
        """
        boveda = await self.get_state(EstadoBoveda)
        if not boveda.lista_tesis:
            await boveda.cargar_tesis()
        
        if not boveda.lista_tesis:
            return rx.toast.warning("No hay Trabajos de Grado registrados para exportar.")
        
        try:
            import io
            import os
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Bóveda de Trabajos de Grado"
            
            # Estilos profesionales
            color_primario = "6366F1" # Indigo
            color_secundario = "F1F5F9"
            
            fuente_titulo = Font(name='Arial', size=16, bold=True, color="FFFFFF")
            fuente_header = Font(name='Arial', size=12, bold=True, color="FFFFFF")
            fuente_normal = Font(name='Arial', size=11)
            
            alineacion_centro = Alignment(horizontal='center', vertical='center')
            alineacion_izq = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            relleno_titulo = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
            relleno_header = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
            relleno_fila_par = PatternFill(start_color=color_secundario, end_color=color_secundario, fill_type="solid")
            
            borde_delgado = Border(
                left=Side(style='thin', color="CBD5E1"),
                right=Side(style='thin', color="CBD5E1"),
                top=Side(style='thin', color="CBD5E1"),
                bottom=Side(style='thin', color="CBD5E1")
            )

            # Ajustar altura de filas de encabezado
            ws.row_dimensions[1].height = 80
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 10  # fila separadora

            # Logo IUTEPI (esquina superior derecha)
            logo_path = os.path.join(os.path.dirname(__file__), "..", "..", "assets", "iutepi.png")
            logo_path = os.path.normpath(logo_path)
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                if img.height > 0:
                    ratio = img.width / img.height
                    img.height = 75
                    img.width = int(ratio * 75)
                ws.add_image(img, "J1")

            # Encabezado del reporte
            ws.merge_cells('A1:I1')
            ws['A1'] = "IUTEPI - REPORTE DE BÓVEDA DE TRABAJOS DE GRADO"
            ws['A1'].font = fuente_titulo
            ws['A1'].alignment = alineacion_centro
            ws['A1'].fill = relleno_titulo
            # Aplicar el mismo fondo en la celda del logo
            ws['J1'].fill = relleno_titulo

            ws.merge_cells('A2:J2')
            ws['A2'] = f"Descripción: Listado detallado de todos los Trabajos de Grado registrados en el sistema. | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].font = Font(name='Arial', size=11, italic=True)
            ws['A2'].alignment = alineacion_centro
            
            # Cabeceras de tabla
            headers = ["ID", "Cédula", "Nombre", "Apellido", "Carrera", "Título", "Tutor Académico", "Tutor Empresarial", "Empresa", "Pública"]
            for col_num, header in enumerate(headers, 1):
                celda = ws.cell(row=4, column=col_num, value=header)
                celda.font = fuente_header
                celda.alignment = alineacion_centro
                celda.fill = relleno_header
                celda.border = borde_delgado
            
            ws.freeze_panes = 'A5'
            ws.sheet_view.showGridLines = False
            ultima_letra = ws.cell(row=4, column=len(headers)).column_letter
            ws.auto_filter.ref = f"A4:{ultima_letra}{len(boveda.lista_tesis) + 4}"
            
            # Datos
            for row_num, t in enumerate(boveda.lista_tesis, 5):
                ws.row_dimensions[row_num].height = 20
                datos_fila = [
                    t.get("id", ""), t.get("cedula_estudiante", ""), t.get("nombre_estudiante", ""),
                    t.get("apellido_estudiante", ""), t.get("carrera", ""), t.get("titulo", ""),
                    t.get("tutor_academico", ""), t.get("tutor_empresa", ""), t.get("nombre_empresa", ""),
                    "Sí" if t.get("publico", False) else "No",
                ]
                for col_num, valor in enumerate(datos_fila, 1):
                    celda = ws.cell(row=row_num, column=col_num, value=valor)
                    celda.font = fuente_normal
                    celda.alignment = alineacion_izq if col_num > 2 else alineacion_centro
                    celda.border = borde_delgado
                    if row_num % 2 == 0:
                        celda.fill = relleno_fila_par
            
            # Ajustar anchos de columna
            anchos = [5, 12, 15, 15, 20, 40, 20, 20, 20, 10]
            for i, ancho in enumerate(anchos, 1):
                ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = ancho

            salida = io.BytesIO()
            wb.save(salida)
            return rx.download(
                data=salida.getvalue(),
                filename=f"reporte_boveda_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
            )
        except Exception as e:
            logger.exception("Error al generar reporte de bóveda: %s", e)
            return rx.toast.error(f"Error al generar reporte: {e}")

    async def cargar_reportes(self):
        """Carga todos los datos analíticos desde la base de datos."""
        def _fetch_reportes():
            conn = obtener_conexion()
            if conn is None:
                return None, "Error de conexión al cargar reportes."
            try:
                with conn:
                    with conn.cursor() as cursor:
                        # 1. Resumen Global
                        cursor.execute("SELECT COUNT(*) FROM estudiante WHERE esta_activo = TRUE")
                        total = cursor.fetchone()[0]
                        
                        cursor.execute("SELECT COUNT(*) FROM estudiante WHERE tutor_academico_id IS NOT NULL AND esta_activo = TRUE")
                        con_pasantia = cursor.fetchone()[0]
                        
                        cursor.execute("SELECT COUNT(*) FROM trabajo_de_grado")
                        total_tesis = cursor.fetchone()[0]

                        resumen = {
                            "total_estudiantes": total,
                            "con_pasantia": con_pasantia,
                            "sin_pasantia": total - con_pasantia,
                            "total_tesis": total_tesis
                        }

                        # 2. Estadísticas por Carrera
                        cursor.execute("""
                            SELECT c.nombre, COUNT(e.id) 
                            FROM carrera c 
                            LEFT JOIN estudiante e ON c.id = e.carrera_id AND e.esta_activo = TRUE
                            WHERE c.esta_activa = TRUE
                            GROUP BY c.nombre
                        """)
                        resultados_carreras = cursor.fetchall()
                        max_est = max([r[1] for r in resultados_carreras]) if resultados_carreras else 1
                        estadisticas_carreras = [
                            {"nombre": r[0], "cantidad": r[1], "progreso": (r[1] / max_est) * 100}
                            for r in resultados_carreras
                        ]

                        # 3. Mejores Tutores (Top 5)
                        cursor.execute("""
                            SELECT ta.nombre || ' ' || ta.apellido, COUNT(e.id) as cantidad
                            FROM tutor_academico ta
                            JOIN estudiante e ON ta.id = e.tutor_academico_id
                            WHERE ta.esta_activo = TRUE
                            GROUP BY ta.id, ta.nombre, ta.apellido
                            ORDER BY cantidad DESC LIMIT 5
                        """)
                        mejores_tutores = [{"nombre": r[0], "cantidad": r[1]} for r in cursor.fetchall()]

                        # 4. Todas las Empresas (con conteo de pasantes)
                        # Correo y teléfono se toman del tutor empresarial asociado
                        cursor.execute("""
                            SELECT emp.nombre, emp.direccion, MAX(te.correo), MAX(te.telefono),
                                   COUNT(e.id) as cantidad, MAX(te.nombre) as tutor_nombre
                            FROM empresa emp
                            LEFT JOIN tutor_empresarial te ON emp.id = te.empresa_id
                            LEFT JOIN estudiante e ON te.id = e.tutor_empresarial_id AND e.esta_activo = TRUE
                            GROUP BY emp.id, emp.nombre, emp.direccion
                            ORDER BY cantidad DESC, emp.nombre ASC
                        """)
                        mejores_empresas = [
                            {"nombre": r[0], "direccion": r[1], "correo": r[2], "telefono": r[3],
                             "cantidad": r[4], "tutor": r[5] or "No asignado"}
                            for r in cursor.fetchall()
                        ]
                return {
                    "resumen": resumen,
                    "estadisticas_carreras": estadisticas_carreras,
                    "mejores_tutores": mejores_tutores,
                    "mejores_empresas": mejores_empresas,
                }, None
            except Exception as e:
                logger.exception("Error al cargar reportes: %s", e)
                return None, f"Error al cargar reportes: {e}"
            finally:
                if conn:
                    try:
                        conn.close()
                    except Exception:
                        pass

        resultado, error = await asyncio.to_thread(_fetch_reportes)
        if error:
            return rx.toast.error(error)

        if resultado is None:
            return rx.toast.error("Error desconocido al cargar reportes.")

        self.resumen_global = resultado["resumen"]
        self.estadisticas_carreras = resultado["estadisticas_carreras"]
        self.mejores_tutores = resultado["mejores_tutores"]
        self.mejores_empresas = resultado["mejores_empresas"]

    def exportar_empresas_excel(self):
        """Genera un reporte Excel profesional de todas las empresas vinculadas."""
        if not self.mejores_empresas:
            return rx.toast.warning("No hay datos de empresas para exportar.")
            
        try:
            import io
            import os
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Vinculación Empresarial"
            
            # Estilos
            color_primario = "10B981" # Emerald
            color_secundario = "F1F5F9"
            
            fuente_titulo = Font(name='Arial', size=16, bold=True, color="FFFFFF")
            fuente_header = Font(name='Arial', size=12, bold=True, color="FFFFFF")
            fuente_normal = Font(name='Arial', size=11)
            
            alineacion_centro = Alignment(horizontal='center', vertical='center')
            alineacion_izq = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            relleno_titulo = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
            relleno_header = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
            relleno_fila_par = PatternFill(start_color=color_secundario, end_color=color_secundario, fill_type="solid")
            
            borde_delgado = Border(
                left=Side(style='thin', color="CBD5E1"),
                right=Side(style='thin', color="CBD5E1"),
                top=Side(style='thin', color="CBD5E1"),
                bottom=Side(style='thin', color="CBD5E1")
            )

            # Ajustar altura de filas de encabezado
            ws.row_dimensions[1].height = 80
            ws.row_dimensions[2].height = 20

            # Logo IUTEPI (esquina superior derecha)
            logo_path = os.path.join(os.path.dirname(__file__), "..", "..", "assets", "iutepi.png")
            logo_path = os.path.normpath(logo_path)
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                if img.height > 0:
                    ratio = img.width / img.height
                    img.height = 75
                    img.width = int(ratio * 75)
                ws.add_image(img, "F1")

            # Encabezado (A1:E1, logo en F1)
            ws.merge_cells('A1:E1')
            ws['A1'] = "IUTEPI - REPORTE DE VINCULACIÓN EMPRESARIAL"
            ws['A1'].font = fuente_titulo
            ws['A1'].alignment = alineacion_centro
            ws['A1'].fill = relleno_titulo
            ws['F1'].fill = relleno_titulo

            ws.merge_cells('A2:F2')
            ws['A2'] = f"Descripción: Listado detallado de todas las entidades aliadas y su carga de pasantes. | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].font = Font(name='Arial', size=11, italic=True)
            ws['A2'].alignment = alineacion_centro

            headers = ["Empresa", "Dirección", "Tutor Empresarial", "Correo de Contacto", "Teléfono", "Pasantes"]
            for col_num, header in enumerate(headers, 1):
                celda = ws.cell(row=4, column=col_num, value=header)
                celda.font = fuente_header
                celda.alignment = alineacion_centro
                celda.fill = relleno_header
                celda.border = borde_delgado

            ws.freeze_panes = 'A5'
            ws.sheet_view.showGridLines = False
            ultima_letra = ws.cell(row=4, column=len(headers)).column_letter
            ws.auto_filter.ref = f"A4:{ultima_letra}{len(self.mejores_empresas) + 4}"

            for row_num, emp in enumerate(self.mejores_empresas, 5):
                ws.row_dimensions[row_num].height = 20
                datos_fila = [
                    emp["nombre"], emp.get("direccion", ""), emp.get("tutor", "No asignado"),
                    emp.get("correo") or "N/A", emp.get("telefono") or "N/A", emp["cantidad"]
                ]
                for col_num, valor in enumerate(datos_fila, 1):
                    celda = ws.cell(row=row_num, column=col_num, value=valor)
                    celda.font = fuente_normal
                    celda.alignment = alineacion_centro if col_num == 6 else alineacion_izq
                    celda.border = borde_delgado
                    if row_num % 2 == 0:
                        celda.fill = relleno_fila_par

            anchos = [32, 32, 28, 32, 18, 12]
            for i, ancho in enumerate(anchos, 1):
                ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = ancho

            salida = io.BytesIO()
            wb.save(salida)
            return rx.download(
                data=salida.getvalue(),
                filename=f"reporte_empresas_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
            )
        except Exception as e:
            logger.exception("Error al generar reporte excel: %s", e)
            return rx.toast.error("Error al generar el reporte de empresas en Excel.")

    def exportar_empresas_pdf(self):
        """Genera un reporte PDF profesional de las empresas."""
        if not self.mejores_empresas:
            return rx.toast.warning("No hay datos para generar el PDF.")

        try:
            import os
            pdf = FPDF(orientation='L', unit='mm', format='A4')
            pdf.add_page()

            # Logo IUTEPI (esquina superior derecha)
            logo_path = os.path.join(os.path.dirname(__file__), "..", "..", "assets", "iutepi.png")
            logo_path = os.path.normpath(logo_path)
            if os.path.exists(logo_path):
                pdf.image(logo_path, x=240, y=7, w=45, h=22)

            # Encabezado con Identidad Institucional
            pdf.set_xy(10, 10)
            pdf.set_font("Helvetica", 'B', 15)
            pdf.set_text_color(31, 41, 55)  # Gris oscuro
            pdf.cell(140, 7, "Instituto Universitario de Tecnología", ln=True, align='L')
            pdf.cell(140, 7, "para la Informática - IUTEPI", ln=False, align='L')
            pdf.ln(10)

            # Línea decorativa
            pdf.set_draw_color(99, 102, 241)  # Indigo
            pdf.set_line_width(0.7)
            pdf.line(10, pdf.get_y(), 287, pdf.get_y())
            pdf.ln(5)

            # Título del Reporte y Metadatos
            pdf.set_font("Helvetica", 'B', 14)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(0, 10, "Reporte de Vinculación Empresarial", ln=True, align='L')
            
            pdf.set_font("Helvetica", 'I', 9)
            pdf.set_text_color(100, 116, 139)  # Slate
            pdf.cell(0, 5, f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align='L')
            pdf.ln(8)
            
            # Descripción del Reporte
            pdf.set_font("Helvetica", '', 10)
            pdf.set_text_color(30, 41, 59)
            desc = "Listado oficial de entidades receptoras de pasantes y vinculación académica institucional."
            pdf.cell(0, 10, desc, ln=True)
            
            # Resumen de métricas en el PDF
            pdf.set_font("Helvetica", 'B', 10)
            pdf.cell(0, 8, f"Total de Empresas Registradas: {len(self.mejores_empresas)}", ln=True)
            pdf.ln(10)

            # Cabecera de tabla
            pdf.set_fill_color(16, 185, 129)  # Emerald del sistema
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", 'B', 9)

            pdf.cell(50, 10, "Empresa", 1, 0, 'C', True)
            pdf.cell(60, 10, "Dirección", 1, 0, 'C', True)
            pdf.cell(45, 10, "Tutor Empresarial", 1, 0, 'C', True)
            pdf.cell(60, 10, "Correo de Contacto", 1, 0, 'C', True)
            pdf.cell(40, 10, "Teléfono", 1, 0, 'C', True)
            pdf.cell(20, 10, "Pasantes", 1, 1, 'C', True)

            # Datos de la tabla
            pdf.set_text_color(30, 41, 59)
            pdf.set_font("Helvetica", '', 8)

            relleno = False
            for emp in self.mejores_empresas:
                pdf.set_fill_color(248, 250, 252) if relleno else pdf.set_fill_color(255, 255, 255)

                nombre = (emp["nombre"][:28] + '..') if len(emp["nombre"]) > 30 else emp["nombre"]
                direccion = (emp.get("direccion", "")[:32] + '..') if len(emp.get("direccion", "")) > 34 else emp.get("direccion", "")
                tutor_val = emp.get("tutor") or "No asignado"
                tutor_val = (tutor_val[:22] + '..') if len(tutor_val) > 24 else tutor_val
                correo_val = emp.get("correo") or "N/A"
                correo_val = (correo_val[:32] + '..') if len(correo_val) > 34 else correo_val

                pdf.cell(50, 8, nombre, 1, 0, 'L', relleno)
                pdf.cell(60, 8, direccion, 1, 0, 'L', relleno)
                pdf.cell(45, 8, tutor_val, 1, 0, 'L', relleno)
                pdf.cell(60, 8, correo_val, 1, 0, 'L', relleno)
                pdf.cell(40, 8, emp.get("telefono") or "N/A", 1, 0, 'C', relleno)
                pdf.cell(20, 8, str(emp["cantidad"]), 1, 1, 'C', relleno)
                relleno = not relleno

            # Pie de página
            pdf.set_y(-20)
            pdf.set_font("Helvetica", 'I', 8)
            pdf.set_text_color(148, 163, 184)
            pdf.cell(0, 10, f"Página {pdf.page_no()} - Documento Administrativo Confidencial - IUTEPI", align='C')
            
            # Obtener el contenido de forma segura
            pdf_output = bytes(pdf.output())
            
            return rx.download(data=pdf_output, filename=f"reporte_empresas_{datetime.now().strftime('%d_%m_%Y')}.pdf")
        except Exception as e:
            logger.exception("Error al generar PDF de reportes: %s", e)
            return rx.toast.error("Error técnico al generar el PDF.")

    async def exportar_tesis_pdf(self):
        """Genera un reporte PDF profesional de la bóveda de Trabajos de Grado (horizontal)."""
        boveda = await self.get_state(EstadoBoveda)
        if not boveda.lista_tesis:
            await boveda.cargar_tesis()

        if not boveda.lista_tesis:
            return rx.toast.warning("No hay tesis registradas para exportar.")

        try:
            import os
            pdf = FPDF(orientation='L', unit='mm', format='A4')  # Apaisado para más columnas
            pdf.add_page()

            # Logo IUTEPI (esquina superior derecha)
            logo_path = os.path.join(os.path.dirname(__file__), "..", "..", "assets", "iutepi.png")
            logo_path = os.path.normpath(logo_path)
            if os.path.exists(logo_path):
                pdf.image(logo_path, x=240, y=7, w=45, h=22)

            # Encabezado con Identidad Institucional
            pdf.set_xy(10, 10)
            pdf.set_font("Helvetica", 'B', 15)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(140, 7, "Instituto Universitario de Tecnología", ln=True, align='L')
            pdf.cell(140, 7, "para la Informática - IUTEPI", ln=False, align='L')
            pdf.ln(10)

            pdf.set_draw_color(99, 102, 241)
            pdf.set_line_width(0.7)
            pdf.line(10, pdf.get_y(), 287, pdf.get_y())
            pdf.ln(5)

            pdf.set_font("Helvetica", 'B', 13)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(0, 8, "Reporte de Bóveda de Trabajos de Grado", ln=True, align='L')

            pdf.set_font("Helvetica", 'I', 9)
            pdf.set_text_color(100, 116, 139)
            pdf.cell(0, 5, f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total registros: {len(boveda.lista_tesis)}", ln=True, align='L')
            pdf.ln(6)

            # Cabecera de tabla
            pdf.set_fill_color(99, 102, 241)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", 'B', 8)

            pdf.cell(8, 10, "ID",        1, 0, 'C', True)
            pdf.cell(20, 10, "Cédula",    1, 0, 'C', True)
            pdf.cell(25, 10, "Nombre",    1, 0, 'C', True)
            pdf.cell(25, 10, "Apellido",  1, 0, 'C', True)
            pdf.cell(30, 10, "Carrera",   1, 0, 'C', True)
            pdf.cell(45, 10, "Título",    1, 0, 'C', True)
            pdf.cell(35, 10, "Tutor Académico", 1, 0, 'C', True)
            pdf.cell(35, 10, "Tutor Emp.", 1, 0, 'C', True)
            pdf.cell(35, 10, "Empresa",   1, 0, 'C', True)
            pdf.cell(15, 10, "Público",   1, 1, 'C', True)

            # Filas de datos
            pdf.set_text_color(30, 41, 59)
            pdf.set_font("Helvetica", '', 7)
            relleno = False
            for t in boveda.lista_tesis:
                pdf.set_fill_color(248, 250, 252) if relleno else pdf.set_fill_color(255, 255, 255)

                titulo_t = t.get("titulo", "")
                titulo_t = (titulo_t[:28] + '..') if len(titulo_t) > 30 else titulo_t
                tutor_acad_t = t.get("tutor_academico", "") or ""
                tutor_acad_t = (tutor_acad_t[:22] + '..') if len(tutor_acad_t) > 24 else tutor_acad_t
                tutor_emp_t = t.get("tutor_empresa", "") or ""
                tutor_emp_t = (tutor_emp_t[:22] + '..') if len(tutor_emp_t) > 24 else tutor_emp_t
                empresa_t = t.get("nombre_empresa", "") or ""
                empresa_t = (empresa_t[:22] + '..') if len(empresa_t) > 24 else empresa_t

                pdf.cell(8, 7, str(t.get("id", "")),                  1, 0, 'C', relleno)
                pdf.cell(20, 7, t.get("cedula_estudiante", ""),          1, 0, 'C', relleno)
                pdf.cell(25, 7, t.get("nombre_estudiante", "")[:15],     1, 0, 'L', relleno)
                pdf.cell(25, 7, t.get("apellido_estudiante", "")[:15],   1, 0, 'L', relleno)
                pdf.cell(30, 7, t.get("carrera", "")[:18],              1, 0, 'L', relleno)
                pdf.cell(45, 7, titulo_t,                                 1, 0, 'L', relleno)
                pdf.cell(35, 7, tutor_acad_t,                                  1, 0, 'L', relleno)
                pdf.cell(35, 7, tutor_emp_t,                              1, 0, 'L', relleno)
                pdf.cell(35, 7, empresa_t,                                1, 0, 'L', relleno)
                pdf.cell(15, 7, "Sí" if t.get("publico") else "No",      1, 1, 'C', relleno)
                relleno = not relleno

            # Pie de página
            pdf.set_y(-18)
            pdf.set_font("Helvetica", 'I', 8)
            pdf.set_text_color(148, 163, 184)
            pdf.cell(0, 8, f"Página {pdf.page_no()} - Documento Administrativo Confidencial - IUTEPI", align='C')

            pdf_output = bytes(pdf.output())
            return rx.download(
                data=pdf_output,
                filename=f"reporte_boveda_{datetime.now().strftime('%d_%m_%Y')}.pdf"
            )
        except Exception as e:
            logger.exception("Error al generar PDF de bóveda: %s", e)
            return rx.toast.error(f"Error al generar el PDF de bóveda: {e}")