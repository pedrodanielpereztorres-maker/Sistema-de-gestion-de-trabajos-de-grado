import asyncio
import logging
import reflex as rx
from datetime import date, timedelta
from pydantic import BaseModel
from psycopg2 import sql
from .estado_autenticacion import EstadoAutenticacion, EncriptadorContrasena
from typing import List, Dict, Any
from ..database_manager import obtener_conexion
import math

logger = logging.getLogger(__name__)


class CountProxy:
    """Contenedor simple para exponer to_string() usado por la UI.

    Definido a nivel de módulo para evitar forward-references en las
    anotaciones de tipos que Reflex requiere sean resolubles en import.
    """

    def __init__(self, value: int):
        self._value = value

    def to_string(self) -> str:
        return str(self._value)


class EstudianteResumen(BaseModel):
    cedula: str = ""
    nombre: str = ""
    apellido: str = ""
    correo: str = ""
    carrera: str = ""
    carrera_tutor: str = ""
    telefono_personal: str = ""
    periodo_inicio: str = ""
    periodo_cierre: str = ""
    fecha_inicio_formateada: str = ""
    fecha_cierre_formateada: str = ""
    nombre_tutor: str = ""
    tutor_empresa: str = ""
    nombre_empresa: str = ""
    direccion_empresa: str = ""
    correo_empresa: str = ""
    telefono_empresa: str = ""
    inicial: str = ""


class EstadoEstudiante(rx.State):
    cedula: str = ""
    nombre: str = ""
    apellido: str = ""
    correo: str = ""
    carrera: str = ""
    telefono_personal: str = ""
    periodo_inicio: str = date.today().strftime("%Y-%m-%d")
    periodo_cierre: str = (date.today() + timedelta(weeks=12)).strftime("%Y-%m-%d")
    nombre_tutor: str = ""
    tutor_empresa: str = ""
    nombre_empresa: str = ""
    direccion_empresa: str = ""
    correo_empresa: str = ""
    telefono_empresa: str = ""
    haciendo_trabajo_de_grado: bool = False
    tutor_academico_seleccionado: str = ""
    tutores_disponibles: list[str] = []
    tutores_mapeo: list[dict] = []

    filtro_carrera: str = ""
    filtro_estado: str = "Todos"
    busqueda_dinamica: str = ""

    mostrar_modal: bool = False
    en_edicion: bool = False
    lista_estudiantes: List[EstudianteResumen] = []
    carreras_disponibles: list[str] = []
    usuario_encontrado: bool = False
    mostrar_ficha: bool = False
    estudiante_seleccionado: dict = {}

    @rx.var
    def opciones_carreras(self) -> list[str]:
        return ["Todas las carreras"] + self.carreras_disponibles

    # Paginación de estudiantes: 10 registros por página
    pagina_actual: int = 1
    registros_por_pagina: int = 10
    total_registros: int = 0
    total_paginas: int = 1

    # Estadísticas por carrera en memoria
    carreras_con_cantidad: list[dict] = []

    # Seguridad en eliminación
    mostrar_modal_confirmacion: bool = False
    password_confirmacion: str = ""
    cedula_a_eliminar: str = ""
    # Modal para mostrar ficha de empresa
    empresa_modal_visible: bool = False

    async def cargar_estudiantes(self) -> None:
        auth_state = await self.get_state(EstadoAutenticacion)
        if auth_state.rol_usuario != "administrador":
            return

        condiciones = [sql.SQL("e.esta_activo = TRUE")]
        params: list = []

        if self.busqueda_dinamica:
            filtro = f"%{self.busqueda_dinamica.strip().lower()}%"
            condiciones.append(
                sql.SQL(
                    "(LOWER(e.cedula) LIKE %s OR LOWER(e.nombre) LIKE %s OR LOWER(e.apellido) LIKE %s OR LOWER(u.correo) LIKE %s)"
                )
            )
            params.extend([filtro, filtro, filtro, filtro])

        if self.filtro_carrera and self.filtro_carrera != "Todas las carreras":
            condiciones.append(sql.SQL("c.nombre = %s"))
            params.append(self.filtro_carrera)

        if self.filtro_estado == "En Pasantía":
            condiciones.append(sql.SQL("e.tutor_academico_id IS NOT NULL"))
        elif self.filtro_estado == "Sin Pasantía":
            condiciones.append(sql.SQL("e.tutor_academico_id IS NULL"))

        where_clause = sql.SQL(" AND ").join(condiciones)

        query_total = sql.SQL(
            """
            SELECT COUNT(*) FROM estudiante e
            LEFT JOIN usuario u ON e.usuario_id = u.id
            LEFT JOIN carrera c ON e.carrera_id = c.id
            WHERE {where}
            """
        ).format(where=where_clause)

        query_estudiantes = sql.SQL(
            """
            SELECT 
                e.cedula, e.nombre, e.apellido, c.nombre as carrera_nom,
                e.celular, e.periodo_inicio, e.periodo_cierre,
                ta.nombre || ' ' || ta.apellido as tutor_acad,
                tc.nombre as tutor_carrera,
                te.nombre as tutor_emp, emp.nombre as empresa_nom,
                emp.direccion, te.correo, te.telefono, u.correo
            FROM estudiante e
            JOIN carrera c ON e.carrera_id = c.id
            LEFT JOIN usuario u ON e.usuario_id = u.id
            LEFT JOIN tutor_academico ta ON e.tutor_academico_id = ta.id
            LEFT JOIN carrera tc ON ta.carrera_id = tc.id
            LEFT JOIN tutor_empresarial te ON e.tutor_empresarial_id = te.id
            LEFT JOIN empresa emp ON te.empresa_id = emp.id
            WHERE {where}
            ORDER BY e.id
            LIMIT %s OFFSET %s
            """
        ).format(where=where_clause)

        def _fetch_total():
            conn2 = obtener_conexion()
            if conn2 is None:
                return 0
            try:
                with conn2:
                    with conn2.cursor() as cursor:
                        cursor.execute(query_total, tuple(params))
                        return cursor.fetchone()[0] or 0
            except Exception as exc:
                logger.exception("Error al contar estudiantes: %s", exc)
                return 0
            finally:
                try:
                    conn2.close()
                except Exception:
                    pass

        def _fetch_estudiantes(params, limit, offset):
            conn2 = obtener_conexion()
            if conn2 is None:
                return []
            try:
                with conn2:
                    with conn2.cursor() as cur:
                        cur.execute(query_estudiantes, tuple(params) + (limit, offset))
                        return cur.fetchall()
            except Exception as exc:
                logger.exception("Error al obtener estudiantes: %s", exc)
                return []
            finally:
                try:
                    conn2.close()
                except Exception:
                    pass

        def _fetch_carreras():
            conn3 = obtener_conexion()
            if conn3 is None:
                return []
            try:
                with conn3:
                    with conn3.cursor() as cur:
                        cur.execute(
                            "SELECT nombre FROM carrera WHERE esta_activa = TRUE ORDER BY nombre"
                        )
                        return [rr[0] for rr in cur.fetchall()]
            except Exception as exc:
                logger.exception("Error al obtener carreras disponibles: %s", exc)
                return []
            finally:
                try:
                    conn3.close()
                except Exception:
                    pass

        try:
            total = await asyncio.to_thread(_fetch_total)
            self.total_registros = total
            self.total_paginas = max(1, math.ceil(total / self.registros_por_pagina))

            if self.pagina_actual > self.total_paginas:
                self.pagina_actual = self.total_paginas
            offset = (self.pagina_actual - 1) * self.registros_por_pagina

            rows = await asyncio.to_thread(
                _fetch_estudiantes,
                params,
                self.registros_por_pagina,
                offset,
            )
            self.carreras_disponibles = await asyncio.to_thread(_fetch_carreras)

            self.lista_estudiantes = [
                EstudianteResumen(
                    cedula=r[0] or "",
                    nombre=r[1] or "",
                    apellido=r[2] or "",
                    correo=r[14] or "",
                    carrera=r[3] or "",
                    telefono_personal=r[4] or "",
                    periodo_inicio=str(r[5]) if r[5] else "",
                    periodo_cierre=str(r[6]) if r[6] else "",
                    fecha_inicio_formateada=(r[5].strftime("%d/%m/%Y") if r[5] else ""),
                    fecha_cierre_formateada=(r[6].strftime("%d/%m/%Y") if r[6] else ""),
                    nombre_tutor=r[7] or "",
                    carrera_tutor=r[8] or "",
                    tutor_empresa=r[9] or "",
                    nombre_empresa=r[10] or "",
                    direccion_empresa=r[11] or "",
                    correo_empresa=r[12] or "",
                    telefono_empresa=r[13] or "",
                    inicial=(r[1][0] if r[1] else ""),
                )
                for r in rows
            ]

            # Cargar estadísticas por carrera en memoria
            await self.cargar_estadisticas_carrera()
        except Exception as e:
            logger.error("Error al cargar estudiantes: %s", e, exc_info=True)

    async def cargar_estadisticas_carrera(self) -> None:
        """Llena carreras_con_cantidad usando consultas SQL y los datos en memoria."""
        conn = None

        def _fetch_carreras_db():
            conn = obtener_conexion()
            if conn is None:
                return []
            try:
                with conn:
                    with conn.cursor() as cursor:
                        cursor.execute(
                            "SELECT nombre FROM carrera WHERE esta_activa = TRUE;"
                        )
                        return [c[0] for c in cursor.fetchall()]
            except Exception as exc:
                logger.exception("Error al obtener carreras para estadísticas: %s", exc)
                return []
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        try:
            carreras_db = await asyncio.to_thread(_fetch_carreras_db)
            conteo = {}
            for e in self.lista_estudiantes:
                conteo[e.carrera] = conteo.get(e.carrera, 0) + 1

            self.carreras_con_cantidad = [
                {
                    "nombre": c,
                    "cantidad": conteo.get(c, 0),
                    "progreso": min(conteo.get(c, 0) * 10, 100),
                }
                for c in carreras_db
            ]
        except Exception as e:
            logger.error(
                "Error al cargar estadísticas de carrera: %s", e, exc_info=True
            )
            self.carreras_con_cantidad = []

    @rx.var
    def estudiantes_filtrados(self) -> List[Dict[str, Any]]:
        lista = list(self.lista_estudiantes or [])
        if self.busqueda_dinamica:
            search = self.busqueda_dinamica.lower()
            lista = [
                e
                for e in lista
                if search in getattr(e, "cedula", "").lower()
                or search in getattr(e, "nombre", "").lower()
                or search in getattr(e, "apellido", "").lower()
            ]
        if (
            self.filtro_carrera
            and self.filtro_carrera != "Todas las carreras"
            and self.filtro_carrera != ""
        ):
            lista = [
                e for e in lista if getattr(e, "carrera", "") == self.filtro_carrera
            ]
        if self.filtro_estado == "En Pasantía":
            lista = [e for e in lista if getattr(e, "nombre_tutor", "")]
        elif self.filtro_estado == "Sin Pasantía":
            lista = [e for e in lista if not getattr(e, "nombre_tutor", "")]

        # Convertir a lista de dicts serializables para Reflex
        def _to_dict(e):
            return {
                "cedula": getattr(e, "cedula", ""),
                "nombre": getattr(e, "nombre", ""),
                "apellido": getattr(e, "apellido", ""),
                "carrera": getattr(e, "carrera", ""),
                "telefono_personal": getattr(e, "telefono_personal", ""),
                "periodo_inicio": getattr(e, "periodo_inicio", ""),
                "periodo_cierre": getattr(e, "periodo_cierre", ""),
                "fecha_inicio_formateada": getattr(e, "fecha_inicio_formateada", ""),
                "fecha_cierre_formateada": getattr(e, "fecha_cierre_formateada", ""),
                "nombre_tutor": getattr(e, "nombre_tutor", ""),
                "carrera_tutor": getattr(e, "carrera_tutor", ""),
                "tutor_empresa": getattr(e, "tutor_empresa", ""),
                "nombre_empresa": getattr(e, "nombre_empresa", ""),
                "direccion_empresa": getattr(e, "direccion_empresa", ""),
                "correo_empresa": getattr(e, "correo_empresa", ""),
                "telefono_empresa": getattr(e, "telefono_empresa", ""),
                "inicial": getattr(e, "inicial", ""),
            }

        return [_to_dict(e) for e in lista]

    async def pagina_siguiente(self) -> None:
        if self.pagina_actual < self.total_paginas:
            self.pagina_actual += 1
            await self.cargar_estudiantes()

    async def pagina_anterior(self) -> None:
        if self.pagina_actual > 1:
            self.pagina_actual -= 1
            await self.cargar_estudiantes()

    async def ir_a_pagina(self, n: int) -> None:
        if 1 <= n <= self.total_paginas:
            self.pagina_actual = n
            await self.cargar_estudiantes()

    def abrir_modal(self) -> None:
        self.mostrar_modal = True
        self.en_edicion = False
        self.estudiante_seleccionado = {}
        self.cedula = ""
        self.nombre = ""
        self.apellido = ""
        self.correo = ""
        self.carrera = ""
        self.telefono_personal = ""
        self.periodo_inicio = date.today().strftime("%Y-%m-%d")
        self.periodo_cierre = (date.today() + timedelta(weeks=12)).strftime("%Y-%m-%d")
        self.nombre_tutor = ""
        self.tutor_empresa = ""
        self.nombre_empresa = ""
        self.direccion_empresa = ""
        self.correo_empresa = ""
        self.telefono_empresa = ""
        self.haciendo_trabajo_de_grado = False
        self.tutor_academico_seleccionado = ""
        self.tutores_mapeo = []
        self.usuario_encontrado = False

    async def abrir_modal_edicion_estudiante(self, cedula: str) -> None:
        """Carga los datos del estudiante seleccionado y abre el modal en modo edición."""
        try:
            cedula_str = str(cedula)
        except Exception:
            return
        # Buscar el estudiante en la lista cargada en memoria
        est = next(
            (
                e
                for e in self.lista_estudiantes
                if getattr(e, "cedula", "") == cedula_str
            ),
            None,
        )
        if not est:
            return
        self.cedula = getattr(est, "cedula", "")
        self.nombre = getattr(est, "nombre", "")
        self.apellido = getattr(est, "apellido", "")
        self.correo = getattr(est, "correo", "")
        self.carrera = getattr(est, "carrera", "")
        self.telefono_personal = getattr(est, "telefono_personal", "")
        self.periodo_inicio = getattr(est, "periodo_inicio", "")
        self.periodo_cierre = getattr(est, "periodo_cierre", "")
        self.nombre_tutor = getattr(est, "nombre_tutor", "")
        self.tutor_empresa = getattr(est, "tutor_empresa", "")
        self.nombre_empresa = getattr(est, "nombre_empresa", "")
        self.direccion_empresa = getattr(est, "direccion_empresa", "")
        self.correo_empresa = getattr(est, "correo_empresa", "")
        self.telefono_empresa = getattr(est, "telefono_empresa", "")

        self.haciendo_trabajo_de_grado = bool(self.nombre_tutor)
        if self.carrera:
            await self.cargar_tutores_por_carrera(self.carrera)
            self.tutor_academico_seleccionado = self.nombre_tutor
        else:
            self.tutor_academico_seleccionado = ""

        self.usuario_encontrado = True
        self.en_edicion = True
        self.mostrar_modal = True

    def cerrar_modal(self) -> None:
        self.mostrar_modal = False
        self.en_edicion = False
        self.estudiante_seleccionado = {}

    async def fijar_busqueda_dinamica(self, val: str) -> None:
        """Setter simple para la búsqueda dinámica en la UI. Recarga la lista de estudiantes."""
        try:
            self.busqueda_dinamica = val
        except Exception:
            self.busqueda_dinamica = str(val)
        self.pagina_actual = 1
        await self.cargar_estudiantes()

    async def fijar_filtro_carrera(self, val: str) -> None:
        """Setter para filtro de carrera. Recarga la lista de estudiantes."""
        try:
            self.filtro_carrera = val
        except Exception:
            self.filtro_carrera = str(val)
        self.pagina_actual = 1
        await self.cargar_estudiantes()

    async def limpiar_filtros(self) -> None:
        """Limpia todos los filtros de búsqueda y carrera en la tabla de estudiantes."""
        self.busqueda_dinamica = ""
        self.filtro_carrera = "Todas las carreras"
        self.pagina_actual = 1
        await self.cargar_estudiantes()

    async def cargar_datos_usuario(self) -> None:
        """Carga datos básicos de usuario/estudiante por cédula y los coloca en el estado.
        Método llamado desde el on_blur del campo cédula en la UI.
        """
        ced = (self.cedula or "").strip()
        if not ced:
            self.usuario_encontrado = False
            return

        def _fetch_usuario_y_estudiante(cedula: str):
            conn = obtener_conexion()
            if conn is None:
                return None, None
            try:
                with conn:
                    with conn.cursor() as cursor:
                        cursor.execute(
                            "SELECT nombre, apellido, correo FROM usuario WHERE cedula = %s AND esta_activo = TRUE",
                            (cedula,),
                        )
                        usuario_row = cursor.fetchone()

                        cursor.execute(
                            "SELECT c.nombre, e.celular, e.periodo_inicio, e.periodo_cierre, ta.nombre || ' ' || ta.apellido, te.nombre, emp.nombre, emp.direccion, te.correo, te.telefono FROM estudiante e LEFT JOIN carrera c ON e.carrera_id = c.id LEFT JOIN tutor_academico ta ON e.tutor_academico_id = ta.id LEFT JOIN tutor_empresarial te ON e.tutor_empresarial_id = te.id LEFT JOIN empresa emp ON te.empresa_id = emp.id WHERE e.cedula = %s AND e.esta_activo = TRUE",
                            (cedula,),
                        )
                        estudiante_row = cursor.fetchone()
                return usuario_row, estudiante_row
            except Exception as exc:
                logger.exception(
                    "Error al obtener datos de usuario/estudiante: %s", exc
                )
                return None, None
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        usuario_row, estudiante_row = await asyncio.to_thread(
            _fetch_usuario_y_estudiante, ced
        )
        if usuario_row:
            self.nombre = usuario_row[0] or ""
            self.apellido = usuario_row[1] or ""
            self.correo = usuario_row[2] or ""
            self.usuario_encontrado = True
        else:
            self.nombre = ""
            self.apellido = ""
            self.correo = ""
            self.usuario_encontrado = False

        if estudiante_row:
            self.carrera = estudiante_row[0] or ""
            self.telefono_personal = estudiante_row[1] or ""
            self.periodo_inicio = (
                str(estudiante_row[2]) if estudiante_row[2] else self.periodo_inicio
            )
            self.periodo_cierre = (
                str(estudiante_row[3]) if estudiante_row[3] else self.periodo_cierre
            )
            self.nombre_tutor = estudiante_row[4] or ""
            self.tutor_empresa = estudiante_row[5] or ""
            self.nombre_empresa = estudiante_row[6] or ""
            self.direccion_empresa = estudiante_row[7] or ""
            self.correo_empresa = estudiante_row[8] or ""
            self.telefono_empresa = estudiante_row[9] or ""
            self.en_edicion = True
        else:
            self.en_edicion = False

        if self.carrera:
            await self.cargar_tutores_por_carrera(self.carrera)
            self.haciendo_trabajo_de_grado = bool(self.nombre_tutor)
            self.tutor_academico_seleccionado = self.nombre_tutor
        else:
            self.haciendo_trabajo_de_grado = False
            self.tutor_academico_seleccionado = ""

    async def cargar_tutores_por_carrera(self, carrera: str = None) -> None:
        """Llena self.tutores_disponibles con tutores académicos para la carrera indicada.
        Si carrera es None, usa self.carrera.
        """
        carrera_nombre = (carrera or self.carrera or "").strip()
        if not carrera_nombre:
            self.tutores_disponibles = []
            self.tutores_mapeo = []
            return

        def _fetch_tutores(carrera: str):
            conn = obtener_conexion()
            if conn is None:
                return []
            try:
                with conn:
                    with conn.cursor() as cursor:
                        cursor.execute(
                            "SELECT ta.id, ta.nombre || ' ' || ta.apellido FROM tutor_academico ta JOIN carrera c ON ta.carrera_id = c.id WHERE c.nombre = %s AND ta.esta_activo = TRUE ORDER BY ta.nombre",
                            (carrera,),
                        )
                        return cursor.fetchall()
            except Exception as exc:
                logger.exception("Error al obtener tutores por carrera: %s", exc)
                return []
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        rows = await asyncio.to_thread(_fetch_tutores, carrera_nombre)
        self.tutores_disponibles = [r[1] for r in rows]
        self.tutores_mapeo = [{"id": r[0], "nombre": r[1]} for r in rows]

    def abrir_modal_confirmacion(self, cedula: str) -> None:
        try:
            self.cedula_a_eliminar = cedula
        except Exception:
            self.cedula_a_eliminar = str(cedula)
        self.password_confirmacion = ""
        self.mostrar_modal_confirmacion = True

    def cerrar_modal_confirmacion(self) -> None:
        self.mostrar_modal_confirmacion = False
        self.password_confirmacion = ""
        self.cedula_a_eliminar = ""

    # Métodos para mostrar/ocultar la tarjeta flotante de la empresa
    def abrir_modal_empresa(
        self,
        nombre_empresa: str,
        direccion: str,
        correo: str,
        telefono: str,
        tutor_empresa: str,
    ) -> None:
        try:
            self.nombre_empresa = nombre_empresa or ""
            self.direccion_empresa = direccion or ""
            self.correo_empresa = correo or ""
            self.telefono_empresa = telefono or ""
            self.tutor_empresa = tutor_empresa or ""
        except Exception:
            # Fallback cuando se reciben objetos reactivos
            self.nombre_empresa = str(nombre_empresa)
            self.direccion_empresa = str(direccion)
            self.correo_empresa = str(correo)
            self.telefono_empresa = str(telefono)
            self.tutor_empresa = str(tutor_empresa)
        self.empresa_modal_visible = True

    def cerrar_modal_empresa(self) -> None:
        self.empresa_modal_visible = False

    async def confirmar_eliminacion_estudiante(self) -> rx.Component:
        if not self.cedula_a_eliminar:
            return rx.toast.error("No se seleccionó ningún estudiante.")

        if not self.password_confirmacion:
            return rx.toast.error(
                "Debe ingresar su contraseña para confirmar la eliminación."
            )

        estado_auth = await self.get_state(EstadoAutenticacion)
        if not estado_auth.usuario:
            return rx.toast.error("Sesión no válida o expirada.")

        def _deactivate_student(usuario_id: int, cedula: str, password: str):
            conn = obtener_conexion()
            if conn is None:
                return False, "Error de conexión al servidor."
            try:
                with conn:
                    with conn.cursor() as cursor:
                        cursor.execute(
                            "SELECT contrasena_hash FROM usuario WHERE id = %s AND esta_activo = TRUE",
                            (usuario_id,),
                        )
                        resultado = cursor.fetchone()
                        if not resultado:
                            return False, "Usuario no registrado o inactivo."

                        hash_almacenado = resultado[0]
                        if not EncriptadorContrasena.verificar(
                            password, hash_almacenado
                        ):
                            return False, "La contraseña ingresada es incorrecta."

                        cursor.execute(
                            "UPDATE estudiante SET esta_activo = FALSE WHERE cedula = %s",
                            (cedula,),
                        )
                        cursor.execute(
                            "UPDATE usuario SET esta_activo = FALSE WHERE cedula = %s",
                            (cedula,),
                        )
                    conn.commit()
                return True, ""
            except Exception as exc:
                try:
                    conn.rollback()
                except Exception:
                    pass
                logger.exception("Error al desactivar estudiante: %s", exc)
                return False, f"Error al desactivar: {exc}"
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        ok, mensaje = await asyncio.to_thread(
            _deactivate_student,
            estado_auth.usuario.id,
            self.cedula_a_eliminar,
            self.password_confirmacion,
        )

        if not ok:
            return rx.toast.error(mensaje)

        await self.cargar_estudiantes()
        self.cerrar_modal_confirmacion()
        return rx.toast.success("Estudiante desactivado correctamente.")

    async def guardar_estudiante(self) -> rx.Component:
        """Inserta o actualiza un estudiante y usuario según el estado de edición.
        Valida coherencia de fechas y realiza las operaciones en BD de forma segura.
        """
        if not self.cedula or not self.nombre or not self.apellido or not self.correo:
            return rx.toast.error("Cédula, nombre, apellido y correo son obligatorios.")

        # Validar formato básico del correo
        patron_correo = r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$"
        if not re.match(patron_correo, self.correo.strip()):
            return rx.toast.error(
                "El correo del estudiante no tiene un formato válido."
            )

        # Validación de fechas antes de tocar la BD
        try:
            inicio = date.fromisoformat(self.periodo_inicio)
            cierre = date.fromisoformat(self.periodo_cierre)
        except Exception:
            return rx.toast.error("Formato de fechas inválido. Use AAAA-MM-DD.")

        if cierre <= inicio:
            return rx.toast.error(
                "La fecha de cierre debe ser posterior a la de inicio."
            )

        if inicio.year < date.today().year - 2:
            return rx.toast.error(
                "La fecha de inicio parece incorrecta. Verifique el año."
            )

        duracion_semanas = (cierre - inicio).days / 7
        if duracion_semanas > 52:
            # Advertencia, pero se permite continuar
            # El toast de advertencia no detiene la operación
            rx.toast.warning("La pasantía supera 52 semanas. Verifique la duración.")

        def _save_student():
            conn = obtener_conexion()
            if conn is None:
                return False, "Error de conexión al servidor."
            try:
                with conn:
                    with conn.cursor() as cursor:
                        cursor.execute(
                            "SELECT 1 FROM usuario WHERE LOWER(correo) = LOWER(%s) AND cedula <> %s",
                            (self.correo.strip(), self.cedula),
                        )
                        correo_in_use = cursor.fetchone()
                        if correo_in_use:
                            return (
                                False,
                                "El correo ya está registrado para otro usuario.",
                            )

                        cursor.execute(
                            "SELECT id FROM usuario WHERE cedula = %s", (self.cedula,)
                        )
                        exists = cursor.fetchone()
                        if not exists:
                            from .estado_autenticacion import EncriptadorContrasena

                            hash_clave = EncriptadorContrasena.encriptar(self.cedula)
                            cursor.execute(
                                "SELECT id FROM rol WHERE LOWER(nombre) = 'estudiante'"
                            )
                            rol_row = cursor.fetchone()
                            rol_id = rol_row[0] if rol_row else None
                            cursor.execute(
                                "INSERT INTO usuario (cedula, nombre, apellido, correo, contrasena_hash, rol_id, esta_activo) VALUES (%s, %s, %s, %s, %s, %s, TRUE) RETURNING id",
                                (
                                    self.cedula,
                                    self.nombre,
                                    self.apellido,
                                    self.correo.strip().lower(),
                                    hash_clave,
                                    rol_id,
                                ),
                            )
                            usuario_id = cursor.fetchone()[0]
                        else:
                            usuario_id = exists[0]
                            cursor.execute(
                                "UPDATE usuario SET nombre = %s, apellido = %s, correo = %s WHERE id = %s",
                                (
                                    self.nombre,
                                    self.apellido,
                                    self.correo.strip().lower(),
                                    usuario_id,
                                ),
                            )

                        carrera_id = None
                        if self.carrera:
                            cursor.execute(
                                "SELECT id FROM carrera WHERE nombre = %s",
                                (self.carrera,),
                            )
                            r = cursor.fetchone()
                            if r:
                                carrera_id = r[0]

                        tutor_academico_id = None
                        if self.haciendo_trabajo_de_grado and self.tutor_academico_seleccionado:
                            for t in self.tutores_mapeo:
                                if t["nombre"] == self.tutor_academico_seleccionado:
                                    tutor_academico_id = t["id"]
                                    break
                            if tutor_academico_id is None:
                                cursor.execute(
                                    "SELECT id FROM tutor_academico WHERE nombre || ' ' || apellido = %s AND esta_activo = TRUE",
                                    (self.tutor_academico_seleccionado,),
                                )
                                t_row = cursor.fetchone()
                                if t_row:
                                    tutor_academico_id = t_row[0]

                        tutor_empresarial_id = None
                        if self.nombre_empresa.strip():
                            cursor.execute(
                                "SELECT id FROM empresa WHERE nombre = %s",
                                (self.nombre_empresa.strip(),),
                            )
                            emp_row = cursor.fetchone()
                            if emp_row:
                                empresa_id = emp_row[0]
                                cursor.execute(
                                    "UPDATE empresa SET direccion = %s WHERE id = %s",
                                    (self.direccion_empresa.strip(), empresa_id),
                                )
                            else:
                                cursor.execute(
                                    "INSERT INTO empresa (nombre, direccion) VALUES (%s, %s) RETURNING id",
                                    (
                                        self.nombre_empresa.strip(),
                                        self.direccion_empresa.strip(),
                                    ),
                                )
                                empresa_id = cursor.fetchone()[0]

                            if self.tutor_empresa.strip():
                                cursor.execute(
                                    "SELECT id FROM tutor_empresarial WHERE nombre = %s AND empresa_id = %s",
                                    (self.tutor_empresa.strip(), empresa_id),
                                )
                                te_row = cursor.fetchone()
                                if te_row:
                                    tutor_empresarial_id = te_row[0]
                                    cursor.execute(
                                        "UPDATE tutor_empresarial SET correo = %s, telefono = %s WHERE id = %s",
                                        (
                                            self.correo_empresa.strip(),
                                            self.telefono_empresa.strip(),
                                            tutor_empresarial_id,
                                        ),
                                    )
                                else:
                                    cursor.execute(
                                        "INSERT INTO tutor_empresarial (nombre, correo, telefono, empresa_id) VALUES (%s, %s, %s, %s) RETURNING id",
                                        (
                                            self.tutor_empresa.strip(),
                                            self.correo_empresa.strip(),
                                            self.telefono_empresa.strip(),
                                            empresa_id,
                                        ),
                                    )
                                    tutor_empresarial_id = cursor.fetchone()[0]

                        cursor.execute(
                            "SELECT 1 FROM estudiante WHERE cedula = %s", (self.cedula,)
                        )
                        est_exists = cursor.fetchone()
                        if est_exists:
                            cursor.execute(
                                "UPDATE estudiante SET nombre=%s, apellido=%s, carrera_id=%s, celular=%s, periodo_inicio=%s, periodo_cierre=%s, usuario_id=%s, tutor_academico_id=%s, tutor_empresarial_id=%s WHERE cedula=%s",
                                (
                                    self.nombre,
                                    self.apellido,
                                    carrera_id,
                                    self.telefono_personal,
                                    inicio,
                                    cierre,
                                    usuario_id,
                                    tutor_academico_id,
                                    tutor_empresarial_id,
                                    self.cedula,
                                ),
                            )
                        else:
                            cursor.execute(
                                "INSERT INTO estudiante (cedula, nombre, apellido, carrera_id, celular, periodo_inicio, periodo_cierre, esta_activo, usuario_id, tutor_academico_id, tutor_empresarial_id) VALUES (%s,%s,%s,%s,%s,%s,%s,TRUE,%s,%s,%s)",
                                (
                                    self.cedula,
                                    self.nombre,
                                    self.apellido,
                                    carrera_id,
                                    self.telefono_personal,
                                    inicio,
                                    cierre,
                                    usuario_id,
                                    tutor_academico_id,
                                    tutor_empresarial_id,
                                ),
                            )
                    conn.commit()
                return True, ""
            except Exception as exc:
                try:
                    conn.rollback()
                except Exception:
                    pass
                logger.exception("Error al guardar estudiante: %s", exc)
                return False, str(exc)
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        ok, mensaje = await asyncio.to_thread(_save_student)
        if not ok:
            return rx.toast.error(f"Error al guardar estudiante: {mensaje}")

        await self.cargar_estudiantes()
        self.cerrar_modal()
        return rx.toast.success("Estudiante guardado correctamente.")

    # UI Setters
    def fijar_cedula(self, val: str) -> None:
        try:
            self.cedula = val
        except Exception:
            self.cedula = str(val)

    def fijar_nombre(self, val: str) -> None:
        try:
            self.nombre = val
        except Exception:
            self.nombre = str(val)

    def fijar_apellido(self, val: str) -> None:
        try:
            self.apellido = val
        except Exception:
            self.apellido = str(val)

    def fijar_telefono_personal(self, val: str) -> None:
        try:
            self.telefono_personal = val
        except Exception:
            self.telefono_personal = str(val)

    def fijar_periodo_inicio(self, val: str) -> None:
        try:
            self.periodo_inicio = val
        except Exception:
            self.periodo_inicio = str(val)

    def fijar_periodo_cierre(self, val: str) -> None:
        try:
            self.periodo_cierre = val
        except Exception:
            self.periodo_cierre = str(val)

    def fijar_haciendo_trabajo_de_grado(self, val) -> None:
        try:
            self.haciendo_trabajo_de_grado = bool(val)
        except Exception:
            self.haciendo_trabajo_de_grado = str(val).lower() in ("1", "true", "yes")

    def fijar_tutor_academico_seleccionado(self, val: str) -> None:
        try:
            self.tutor_academico_seleccionado = val
        except Exception:
            self.tutor_academico_seleccionado = str(val)

    def fijar_nombre_empresa(self, val: str) -> None:
        try:
            self.nombre_empresa = val
        except Exception:
            self.nombre_empresa = str(val)

    def fijar_direccion_empresa(self, val: str) -> None:
        try:
            self.direccion_empresa = val
        except Exception:
            self.direccion_empresa = str(val)

    def fijar_tutor_empresa(self, val: str) -> None:
        try:
            self.tutor_empresa = val
        except Exception:
            self.tutor_empresa = str(val)

    def fijar_correo_empresa(self, val: str) -> None:
        try:
            self.correo_empresa = val
        except Exception:
            self.correo_empresa = str(val)

    def fijar_telefono_empresa(self, val: str) -> None:
        try:
            self.telefono_empresa = val
        except Exception:
            self.telefono_empresa = str(val)

    def fijar_carrera(self, val: str) -> None:
        try:
            self.carrera = val
        except Exception:
            self.carrera = str(val)

    def fijar_correo(self, val: str) -> None:
        try:
            self.correo = val
        except Exception:
            self.correo = str(val)

    def fijar_password_confirmacion(self, val: str) -> None:
        try:
            self.password_confirmacion = val
        except Exception:
            self.password_confirmacion = str(val)

    # Compatibilidad con vistas antiguas: contadores y listas esperadas por inicio.py
    @rx.var
    def total_estudiantes(self) -> int:
        """Número total de estudiantes activo."""
        return int(self.total_registros)

    @rx.var
    def estudiantes_en_pasantia(self) -> int:
        c = sum(1 for e in self.lista_estudiantes if getattr(e, "nombre_tutor", ""))
        return int(c)

    @rx.var
    def estudiantes_sin_pasantia(self) -> int:
        c = sum(1 for e in self.lista_estudiantes if not getattr(e, "nombre_tutor", ""))
        return int(c)

    @rx.var
    def lista_con_pasantia(self) -> List[Dict[str, Any]]:
        lista = [
            e for e in (self.lista_estudiantes or []) if getattr(e, "nombre_tutor", "")
        ]
        return [
            {
                "cedula": getattr(e, "cedula", ""),
                "nombre": getattr(e, "nombre", ""),
                "apellido": getattr(e, "apellido", ""),
                "carrera": getattr(e, "carrera", ""),
                "nombre_tutor": getattr(e, "nombre_tutor", ""),
                "nombre_empresa": getattr(e, "nombre_empresa", ""),
                "inicial": getattr(e, "inicial", ""),
            }
            for e in lista
        ]

    @rx.var
    def lista_sin_pasantia(self) -> List[Dict[str, Any]]:
        lista = [
            e
            for e in (self.lista_estudiantes or [])
            if not getattr(e, "nombre_tutor", "")
        ]
        return [
            {
                "cedula": getattr(e, "cedula", ""),
                "nombre": getattr(e, "nombre", ""),
                "apellido": getattr(e, "apellido", ""),
                "carrera": getattr(e, "carrera", ""),
                "inicial": getattr(e, "inicial", ""),
            }
            for e in lista
        ]

    @rx.var
    def estudiantes_por_carrera(self) -> List[Dict[str, Any]]:
        """Lista con nombre, cantidad y progreso para cada carrera (compatibilidad con plantilla)."""
        return self.carreras_con_cantidad

    async def generar_reporte_estudiantes(self) -> rx.Component:
        """Genera y descarga un Excel profesional con todos los estudiantes listados."""
        if not self.lista_estudiantes:
            return rx.toast.warning("No hay estudiantes para exportar.")
        try:
            import io
            import os
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage

            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Estudiantes"

            # Estilos profesionales
            color_primario = "4F46E5"  # Indigo
            color_secundario = "F8FAFC"

            fuente_titulo = Font(name="Arial", size=16, bold=True, color="FFFFFF")
            fuente_header = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            fuente_normal = Font(name="Arial", size=11)

            alineacion_centro = Alignment(horizontal="center", vertical="center")
            alineacion_izq = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )

            relleno_titulo = PatternFill(
                start_color=color_primario, end_color=color_primario, fill_type="solid"
            )
            relleno_header = PatternFill(
                start_color="334155", end_color="334155", fill_type="solid"
            )
            relleno_fila_par = PatternFill(
                start_color=color_secundario,
                end_color=color_secundario,
                fill_type="solid",
            )

            borde_delgado = Border(
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
                top=Side(style="thin", color="E2E8F0"),
                bottom=Side(style="thin", color="E2E8F0"),
            )

            # Ajustar altura de filas de encabezado
            ws.row_dimensions[1].height = 80
            ws.row_dimensions[2].height = 20

            # Logo IUTEPI (esquina superior derecha)
            logo_path = os.path.join(
                os.path.dirname(__file__), "..", "..", "assets", "iutepi.png"
            )
            logo_path = os.path.normpath(logo_path)
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                if img.height > 0:
                    ratio = img.width / img.height
                    img.height = 75
                    img.width = int(ratio * 75)
                ws.add_image(img, "I1")

            # Encabezado
            ws.merge_cells("A1:H1")
            ws["A1"] = "IUTEPI - REPORTE DE ESTUDIANTES EN PASANTÍA"
            ws["A1"].font = fuente_titulo
            ws["A1"].alignment = alineacion_centro
            ws["A1"].fill = relleno_titulo
            ws["I1"].fill = relleno_titulo

            ws.merge_cells("A2:I2")
            ws["A2"] = (
                f"Descripción: Listado detallado del estado actual de todos los alumnos registrados. | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            )
            ws["A2"].font = Font(name="Arial", size=11, italic=True)
            ws["A2"].alignment = alineacion_centro

            headers = [
                "Cédula",
                "Nombre",
                "Apellido",
                "Carrera",
                "Teléfono",
                "Inicio",
                "Cierre",
                "Tutor Académico",
                "Empresa",
            ]
            for col_num, header in enumerate(headers, 1):
                celda = ws.cell(row=4, column=col_num, value=header)
                celda.font = fuente_header
                celda.alignment = alineacion_centro
                celda.fill = relleno_header
                celda.border = borde_delgado

            ws.freeze_panes = "A5"
            ws.sheet_view.showGridLines = False
            ultima_letra = ws.cell(row=4, column=len(headers)).column_letter
            ws.auto_filter.ref = f"A4:{ultima_letra}{len(self.lista_estudiantes) + 4}"

            for row_num, e in enumerate(self.lista_estudiantes, 5):
                ws.row_dimensions[row_num].height = 20
                datos_fila = [
                    getattr(e, "cedula", ""),
                    getattr(e, "nombre", ""),
                    getattr(e, "apellido", ""),
                    getattr(e, "carrera", ""),
                    getattr(e, "telefono_personal", ""),
                    getattr(e, "periodo_inicio", ""),
                    getattr(e, "periodo_cierre", ""),
                    getattr(e, "nombre_tutor", ""),
                    getattr(e, "nombre_empresa", ""),
                ]
                for col_num, valor in enumerate(datos_fila, 1):
                    celda = ws.cell(row=row_num, column=col_num, value=valor)
                    celda.font = fuente_normal
                    celda.alignment = (
                        alineacion_izq
                        if col_num in (2, 3, 4, 8, 9)
                        else alineacion_centro
                    )
                    celda.border = borde_delgado
                    if row_num % 2 == 0:
                        celda.fill = relleno_fila_par

            anchos = [12, 20, 20, 25, 15, 12, 12, 25, 25]
            for i, ancho in enumerate(anchos, 1):
                ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = (
                    ancho
                )

            salida = io.BytesIO()
            wb.save(salida)
            return rx.download(
                data=salida.getvalue(),
                filename=f"reporte_estudiantes_{datetime.now().strftime('%d_%m_%Y')}.xlsx",
            )
        except Exception as e:
            logger.exception("Error al generar reporte de estudiantes: %s", e)
            return rx.toast.error(f"Error al generar reporte: {e}")

    def exportar_estudiantes_pdf(self):
        """Genera un reporte PDF profesional de todos los estudiantes (hoja apaisada)."""
        if not self.lista_estudiantes:
            return rx.toast.warning("No hay estudiantes para exportar.")

        try:
            import os
            from fpdf import FPDF

            pdf = FPDF(orientation="L", unit="mm", format="A4")
            pdf.add_page()

            # Logo IUTEPI (esquina superior derecha)
            logo_path = os.path.join(
                os.path.dirname(__file__), "..", "..", "assets", "iutepi.png"
            )
            logo_path = os.path.normpath(logo_path)
            if os.path.exists(logo_path):
                pdf.image(logo_path, x=240, y=7, w=45, h=22)

            # Encabezado con Identidad Institucional
            pdf.set_xy(10, 10)
            pdf.set_font("Helvetica", "B", 15)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(
                140, 7, "Instituto Universitario de Tecnología", ln=True, align="L"
            )
            pdf.cell(140, 7, "para la Informática - IUTEPI", ln=False, align="L")
            pdf.ln(10)

            pdf.set_draw_color(79, 70, 229)
            pdf.set_line_width(0.7)
            pdf.line(10, pdf.get_y(), 287, pdf.get_y())
            pdf.ln(5)

            pdf.set_font("Helvetica", "B", 13)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(0, 8, "Reporte de Estudiantes en Pasantía", ln=True, align="L")

            pdf.set_font("Helvetica", "I", 9)
            pdf.set_text_color(100, 116, 139)
            pdf.cell(
                0,
                5,
                f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total registros: {len(self.lista_estudiantes)}",
                ln=True,
                align="L",
            )
            pdf.ln(6)

            # Cabecera de tabla
            pdf.set_fill_color(79, 70, 229)  # Indigo
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 8)

            pdf.cell(22, 10, "Cédula", 1, 0, "C", True)
            pdf.cell(30, 10, "Nombre", 1, 0, "C", True)
            pdf.cell(30, 10, "Apellido", 1, 0, "C", True)
            pdf.cell(46, 10, "Carrera", 1, 0, "C", True)
            pdf.cell(25, 10, "Teléfono", 1, 0, "C", True)
            pdf.cell(23, 10, "Inicio", 1, 0, "C", True)
            pdf.cell(23, 10, "Cierre", 1, 0, "C", True)
            pdf.cell(44, 10, "Tutor Académico", 1, 0, "C", True)
            pdf.cell(34, 10, "Empresa", 1, 1, "C", True)

            # Filas de datos
            pdf.set_text_color(30, 41, 59)
            pdf.set_font("Helvetica", "", 7)
            relleno = False
            for e in self.lista_estudiantes:
                (
                    pdf.set_fill_color(248, 250, 252)
                    if relleno
                    else pdf.set_fill_color(255, 255, 255)
                )

                empresa_t = e.get("nombre_empresa", "") or ""
                empresa_t = (
                    (empresa_t[:20] + "..") if len(empresa_t) > 22 else empresa_t
                )
                tutor_t = e.get("nombre_tutor", "") or ""
                tutor_t = (tutor_t[:28] + "..") if len(tutor_t) > 30 else tutor_t

                pdf.cell(22, 7, e.get("cedula", ""), 1, 0, "C", relleno)
                pdf.cell(30, 7, e.get("nombre", "")[:18], 1, 0, "L", relleno)
                pdf.cell(30, 7, e.get("apellido", "")[:18], 1, 0, "L", relleno)
                pdf.cell(46, 7, e.get("carrera", "")[:28], 1, 0, "L", relleno)
                pdf.cell(25, 7, e.get("telefono_personal", ""), 1, 0, "C", relleno)
                pdf.cell(
                    23, 7, e.get("fecha_inicio_formateada", ""), 1, 0, "C", relleno
                )
                pdf.cell(
                    23, 7, e.get("fecha_cierre_formateada", ""), 1, 0, "C", relleno
                )
                pdf.cell(44, 7, tutor_t, 1, 0, "L", relleno)
                pdf.cell(34, 7, empresa_t, 1, 1, "L", relleno)
                relleno = not relleno

            # Pie de página
            pdf.set_y(-18)
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(148, 163, 184)
            pdf.cell(
                0,
                8,
                f"Página {pdf.page_no()} - Documento Administrativo Confidencial - IUTEPI",
                align="C",
            )

            pdf_output = bytes(pdf.output())
            return rx.download(
                data=pdf_output,
                filename=f"reporte_estudiantes_{datetime.now().strftime('%d_%m_%Y')}.pdf",
            )
        except Exception as e:
            logger.exception("Error al generar PDF de estudiantes: %s", e)
            return rx.toast.error(f"Error al generar el PDF: {e}")
